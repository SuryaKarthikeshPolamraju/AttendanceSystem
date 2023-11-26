import openpyxl
from datetime import datetime

def get_next_empty_row(sheet):
    # Find the next empty row in the sheet
    row_index = 1
    while sheet.cell(row=row_index, column=1).value is not None:
        row_index += 1
    return row_index

def update_excel_row(file_path, sheet_name, name, timestamp):
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)

    # Select the sheet by name
    sheet = wb[sheet_name]

    # Get the next empty row
    row_index = get_next_empty_row(sheet)

    # Update the cell values with name and timestamp
    sheet.cell(row=row_index, column=1, value=name)
    sheet.cell(row=row_index, column=2, value=timestamp)

    # Save the changes
    wb.save(file_path)

    print(f"Row {row_index} in sheet '{sheet_name}' updated with Name: {name}, Timestamp: {timestamp}")

# Example usage:
file_path = "data.xlsx"
sheet_name = "Sheet1"
student_name = "John Doe valla tammudu"
timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

update_excel_row(file_path, sheet_name, student_name, timestamp)
